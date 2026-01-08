from typing import Dict, Any, Optional, List, BinaryIO
from src.repository.content_understanding import ContentUnderstandingRepository
from src.repository.storage import MinioStorageRepository, AzureBlobStorageRepository
from src.repository.messaging import RabbitMQRepository, AzureServiceBusRepository
from src.repository.database import AzureCosmosDBRepository
from loguru import logger
from src.domain.file_upload import FileUploadResponse
from src.domain.gl_transaction import GLTransaction
from src.common.const import Environment
from pypdf import PdfReader, PdfWriter
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
import uuid

import os
IS_PRODUCTION = os.getenv("ENV") == Environment.Production.value

# Mapping from XLSX headers to GLTransaction field names (matching Cosmos DB)
XLSX_TO_GL_TRANSACTION_MAP = {
    "CoCd": "cocd",
    "G/L": "gl",
    "Year/month": "year_month",
    "Type": "type",
    "Reference": "reference_number",
    "DocumentNo": "document_number",
    "Supplier": "vendor_code",
    "Supplier Name": "vendor_name",
    "PO Number": "po_number",
    "Tax Based": "tax_based",
    "WHT": "wht",
    "Tax Rate": "tax_rate",
    "URN": "urn",
    "User Name": "username",
    "Text": "text",
    "Clrng doc.": "clearing_document",
    "Clrng doc": "clearing_document",
    "Clearing doc": "clearing_document",
    "Doc. Date": "document_date",
    "Pstng Date": "posting_date",
    "Posting Date": "posting_date",
    "Doc Curr": "document_currency",
    "Doc Currency": "document_currency",
    "Amount in doc. curr.": "amount_in_document_currency",
    "Amount in document currency": "amount_in_document_currency",
    "Loc Curr": "local_currency",
    "Amount in local cur.": "amount_in_local_currency",
    "Ref": "ref",
    "WHT Review": "wht_review",
    "1st Vouching": "first_voucing",
    "2nd Reviewer": "second_reviewer",
    "Type of Tax": "type_of_tax",
    "Docu Ty": "document_type",
}

class GLUpload:
    def __init__(
        self, 
        content_understanding_repo: ContentUnderstandingRepository,
        azure_blob_storage_repo: AzureBlobStorageRepository,
        azure_cosmos_repo: Optional[AzureCosmosDBRepository] = None,
        rabbitmq_repo: Optional[RabbitMQRepository] = None,
        minio_storage_repo: Optional[MinioStorageRepository] = None,
        azure_service_bus_repo: Optional[AzureServiceBusRepository] = None
    ):
        self.content_understanding_repo = content_understanding_repo
        self.azure_blob_storage_repo = azure_blob_storage_repo
        self.azure_cosmos_repo = azure_cosmos_repo
        self.rabbitmq_repo = rabbitmq_repo
        self.minio_storage_repo = minio_storage_repo
        self.azure_service_bus_repo = azure_service_bus_repo

        self.queue_name = "document-uploads"

    def _upload_to_storage(
        self, 
        file: BinaryIO, 
        file_id: str, 
        filename: str, 
        activity_id: str,
        content_type: Optional[str] = None
    ) -> Dict[str, Any]:
        
        storage_repo = self.azure_blob_storage_repo if IS_PRODUCTION else self.minio_storage_repo
        return storage_repo.upload_file(
            file=file,
            file_id=file_id,
            original_filename=filename,
            activity_id=activity_id,
            content_type=content_type
        )
    
    def _upload_single_file(
        self, 
        file_content: bytes, 
        file_id: str, 
        filename: str, 
        activity_id: str,
        page_number: Optional[int] = None,
        content_type: Optional[str] = None
    ) -> Dict[str, Any]:

        file_buffer = BytesIO(file_content)
        file_info = self._upload_to_storage(file_buffer, file_id, filename, activity_id, content_type)
        
        object_name = file_info.get("object_name") or file_info.get("blob_name")
        
        result = {
            "filename": filename,
            "size": file_info.get("size", 0),
            "object_name": object_name
        }
        
        if page_number is not None:
            result["page"] = page_number
            
        return result

    def upload(self, file, file_id: str, original_filename: str) -> Dict[str, Any]:
        
        try:
            file_content = file.read()

            # Extract content_type from original file object
            content_type = getattr(file, 'content_type', 'application/octet-stream')
            
            # Generate activity_id if not provided
            activity_id = str(uuid.uuid4())
            
            # 1. Upload xlsx file to storage
            upload_info = self._upload_single_file(
                file_content=file_content,
                file_id=file_id,
                filename=original_filename,
                activity_id=activity_id,
                content_type=content_type
            )
            
            logger.info(f"File uploaded successfully: {upload_info}")
            
            # 2. Read content by row from xlsx file
            rows_data = self._read_xlsx_rows(file_content)
            
            logger.info(f"Read {len(rows_data)} rows from {original_filename}")
            
            # 3. Insert to Azure Cosmos DB
            if self.azure_cosmos_repo and rows_data:
                try:
                    for row in rows_data:
                        # Convert to GLTransaction model and serialize with aliases (camelCase)
                        gl_transaction = GLTransaction(**row)
                        document_data = gl_transaction.model_dump(by_alias=True)
                        
                        # Ensure glReconItem is present as empty array if None
                        if document_data.get("glReconItem") is None:
                            document_data["glReconItem"] = []
                        
                        # Debug log for critical fields
                        logger.debug(f"Inserting GL transaction with taxRate={document_data.get('taxRate')}, amountInLocalCurrency={document_data.get('amountInLocalCurrency')}")
                        
                        # Insert each GL transaction to Cosmos DB with aliased field names
                        self.azure_cosmos_repo.create_document(
                            container_id="gl-transactions",
                            document_data=document_data
                        )
                except Exception as e:
                    logger.error(f"Error inserting rows to Cosmos DB: {e}")
                    raise
            else:
                logger.warning("Azure Cosmos DB repository not configured or no rows to insert")
            
            # Return as snake_case
            return {
                "file_id": file_id,
                "activity_id": activity_id,
                # "upload_info": upload_info,
                "rows": rows_data,
                "total_rows": len(rows_data)
            }
        
        except Exception as e:
            logger.error(f"Error uploading file {original_filename}: {e}")
            raise
    
    def _read_xlsx_rows(self, file_content: bytes) -> List[Dict[str, Any]]:
        """Read XLSX file and return rows as list of dictionaries matching GLTransaction domain"""
        
        rows = []
        
        try:
            workbook = load_workbook(BytesIO(file_content), data_only=True)
            worksheet = workbook.active
            
            # Get headers from row 2 and strip whitespace
            headers = []
            for cell in worksheet[2]:
                header_value = cell.value.strip() if isinstance(cell.value, str) else cell.value
                headers.append(header_value)
            
            logger.info(f"XLSX headers: {headers}")
            
            # Read data rows starting from row 4
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=4, values_only=False), start=4):
                row_data = {}
                
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx] is not None:
                        xlsx_header = headers[col_idx]
                        # Map XLSX header to GLTransaction field name
                        field_name = XLSX_TO_GL_TRANSACTION_MAP.get(xlsx_header, xlsx_header)
                        # Treat empty strings as None for proper default handling
                        cell_value = cell.value
                        if isinstance(cell_value, str) and cell_value.strip() == "":
                            cell_value = None
                        row_data[field_name] = cell_value
                
                # Only add non-empty rows
                if any(v is not None for v in row_data.values()):
                    # Generate id (Cosmos DB document ID) and gl_transaction_id
                    row_data["id"] = str(uuid.uuid4())
                    row_data["gl_transaction_id"] = f"GLT{datetime.utcnow().strftime('%Y%m%d')}{str(uuid.uuid4())[:8].upper()}"
                    
                    # Set default gl_transaction_status_id
                    row_data["gl_transaction_status_id"] = 1
                    
                    # Initialize gl_recon_item as None (will be converted to [] during insertion)
                    row_data["gl_recon_item"] = None
                    
                    # Set defaults for required fields if missing
                    self._set_default_values(row_data)
                    
                    # Convert numeric values to appropriate types
                    row_data = self._convert_row_types(row_data)
                    
                    rows.append(row_data)
            
            logger.info(f"Successfully read {len(rows)} rows from XLSX file")
            return rows
            
        except Exception as e:
            logger.error(f"Error reading XLSX file: {e}")
            raise
    
    def _set_default_values(self, row_data: Dict[str, Any]) -> None:
        """Set default values for required fields that are missing"""
        
        # Required string fields
        required_string_fields = {
            "urn": "",
            "cocd": "",
            "gl": "",
            "year_month": "",
            "type": "",
            "reference_number": "",
            "document_number": "",
            "po_number": "",
            "username": "",
            "text": "",
            "document_date": "",
            "posting_date": "",
            "document_currency": "IDR",
            "local_currency": "IDR",
            "vendor_id": "",
            "vendor_code": "",
            "vendor_name": "",
            "first_voucing": ""
        }
        
        # Required numeric fields
        required_numeric_fields = {
            "tax_based": 0.0,
            "wht": 0.0,
            "tax_rate": 0.0,
            "amount_in_document_currency": 0.0,
            "amount_in_local_currency": 0.0,
            "wht_normal": 0.0,
            "diff_normal": 0.0
        }
        
        # Set defaults for missing string fields
        for field, default_value in required_string_fields.items():
            if field not in row_data or row_data[field] is None or (isinstance(row_data[field], str) and row_data[field].strip() == ""):
                row_data[field] = default_value
        
        # Set defaults for missing numeric fields
        for field, default_value in required_numeric_fields.items():
            if field not in row_data or row_data[field] is None or (isinstance(row_data[field], str) and row_data[field].strip() == ""):
                row_data[field] = default_value
    
    def _convert_row_types(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Convert row data types to match GLTransaction domain (Cosmos DB)"""
        
        # Fields that should be float
        float_fields = [
            "tax_based", "wht", "tax_rate", 
            "amount_in_document_currency", "amount_in_local_currency",
            "wht_normal", "diff_normal"
        ]
        
        # Fields that should be string (matching Cosmos DB field names)
        string_fields = [
            "id", "cocd", "gl", "year_month", "type", "reference_number",
            "document_number", "vendor_id", "vendor_code", "vendor_name", "po_number", "urn", "username",
            "text", "clearing_document", "document_date", "posting_date",
            "document_currency", "local_currency", "ref", "first_voucing",
            "second_reviewer", "gl_transaction_id"
        ]
        
        # Extra fields from XLSX (not in Cosmos DB schema, kept for reference)
        extra_string_fields = [
            "wht_review", "type_of_tax", "document_type"
        ]
        
        # Convert datetime objects to ISO format strings
        for key, value in row_data.items():
            if isinstance(value, datetime):
                row_data[key] = value.isoformat()
        
        for field in float_fields:
            if field in row_data:
                if row_data[field] is None or (isinstance(row_data[field], str) and row_data[field].strip() == ""):
                    row_data[field] = 0.0
                else:
                    try:
                        row_data[field] = float(row_data[field])
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Failed to convert {field}='{row_data[field]}' to float: {e}. Setting to 0.0")
                        row_data[field] = 0.0
        
        for field in string_fields:
            if field in row_data and row_data[field] is not None:
                row_data[field] = str(row_data[field])
            elif field in row_data:
                row_data[field] = ""
        
        for field in extra_string_fields:
            if field in row_data and row_data[field] is not None:
                row_data[field] = str(row_data[field])
        
        return row_data